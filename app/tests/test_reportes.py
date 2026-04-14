"""Tests for report and export endpoints."""

import io
from datetime import date, time

import pytest
from openpyxl import load_workbook

from app.core.security import create_access_token, hash_password
from app.models.actividad import Actividad
from app.models.asistencia import Asistencia
from app.models.enums import (
    DiaSemana,
    EstadoInscripcion,
    RolUsuario,
)
from app.models.evaluacion_salud import EvaluacionSalud
from app.models.inscripcion import Inscripcion
from app.models.turno import Turno
from app.models.usuario import Usuario


# ── Helpers ──────────────────────────────────────────────────────


_user_counter = 0


async def _create_user(db, *, email, rol=RolUsuario.ALUMNO, nombre="Test", apellido="User"):
    global _user_counter
    _user_counter += 1
    user = Usuario(
        nombre=nombre,
        apellido=apellido,
        email=email,
        password_hash=hash_password("testpassword"),
        rol=rol,
        dni=f"{_user_counter:08d}",
        telefono="555-1234",
        fecha_nacimiento=date(2000, 1, 15),
    )
    db.add(user)
    await db.commit()
    await db.refresh(user)
    return user


async def _create_admin(db):
    return await _create_user(
        db, email="admin@test.com", rol=RolUsuario.ADMIN, nombre="Admin", apellido="Boss"
    )


async def _create_alumno(db, *, email="alumno@test.com", nombre="Alumno", apellido="Test"):
    return await _create_user(db, email=email, rol=RolUsuario.ALUMNO, nombre=nombre, apellido=apellido)


async def _create_profesor(db, *, email="profesor@test.com"):
    return await _create_user(db, email=email, rol=RolUsuario.PROFESOR, nombre="Profesor", apellido="Test")


async def _create_actividad(db, *, nombre="Yoga"):
    act = Actividad(
        nombre=nombre, descripcion="Test activity",
        cupo_maximo=20, duracion_min=60, activa=True,
    )
    db.add(act)
    await db.commit()
    await db.refresh(act)
    return act


async def _create_turno(db, *, actividad_id, profesor_id, dia=DiaSemana.LUNES):
    turno = Turno(
        actividad_id=actividad_id, profesor_id=profesor_id,
        dia_semana=dia, hora_inicio=time(9, 0), hora_fin=time(10, 0),
        sala="Sala A", activo=True,
    )
    db.add(turno)
    await db.commit()
    await db.refresh(turno)
    return turno


async def _create_inscripcion(db, *, alumno_id, turno_id):
    inscripcion = Inscripcion(
        alumno_id=alumno_id, turno_id=turno_id,
        estado=EstadoInscripcion.ACTIVA,
    )
    db.add(inscripcion)
    await db.commit()
    await db.refresh(inscripcion)
    return inscripcion


async def _create_evaluacion(db, *, alumno_id, profesional_id, peso=75.0, imc=24.5, grasa=18.0):
    ev = EvaluacionSalud(
        alumno_id=alumno_id, profesional_id=profesional_id,
        peso_kg=peso, altura_cm=175.0, imc=imc,
        grasa_corporal=grasa, objetivo="Tonificar",
        notas="Buen progreso",
    )
    db.add(ev)
    await db.commit()
    await db.refresh(ev)
    return ev


def _auth(user):
    token = create_access_token(subject=user.id)
    return {"Authorization": f"Bearer {token}"}


def _parse_excel(content: bytes):
    """Load workbook from response bytes."""
    return load_workbook(io.BytesIO(content))


EXCEL_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MEDIA = "application/pdf"


# ── Test: GET /api/reportes/alumnos/excel ────────────────────────


class TestReporteAlumnosExcel:
    @pytest.mark.asyncio
    async def test_admin_downloads_alumnos_excel(self, client, db_session):
        admin = await _create_admin(db_session)
        await _create_alumno(db_session, email="a1@test.com", nombre="Ana", apellido="Garcia")
        await _create_alumno(db_session, email="a2@test.com", nombre="Juan", apellido="Lopez")

        r = await client.get("/api/reportes/alumnos/excel", headers=_auth(admin))
        assert r.status_code == 200
        assert EXCEL_MEDIA in r.headers["content-type"]
        assert "attachment" in r.headers["content-disposition"]
        assert "alumnos.xlsx" in r.headers["content-disposition"]

        wb = _parse_excel(r.content)
        ws = wb.active
        assert ws.title == "Alumnos"
        # Header row + 2 students
        assert ws.max_row == 3
        assert ws.cell(1, 1).value == "ID"
        assert ws.cell(1, 2).value == "Apellido"

    @pytest.mark.asyncio
    async def test_empty_alumnos_returns_headers_only(self, client, db_session):
        admin = await _create_admin(db_session)

        r = await client.get("/api/reportes/alumnos/excel", headers=_auth(admin))
        assert r.status_code == 200
        wb = _parse_excel(r.content)
        ws = wb.active
        assert ws.max_row == 1  # Only headers

    @pytest.mark.asyncio
    async def test_non_admin_forbidden(self, client, db_session):
        alumno = await _create_alumno(db_session)
        r = await client.get("/api/reportes/alumnos/excel", headers=_auth(alumno))
        assert r.status_code == 403

    @pytest.mark.asyncio
    async def test_requires_auth(self, client):
        r = await client.get("/api/reportes/alumnos/excel")
        assert r.status_code in (401, 403)

    @pytest.mark.asyncio
    async def test_styled_headers(self, client, db_session):
        admin = await _create_admin(db_session)
        await _create_alumno(db_session)

        r = await client.get("/api/reportes/alumnos/excel", headers=_auth(admin))
        wb = _parse_excel(r.content)
        ws = wb.active
        header_cell = ws.cell(1, 1)
        assert header_cell.font.bold is True
        assert header_cell.fill.start_color.rgb is not None


# ── Test: GET /api/reportes/asistencias/excel ────────────────────


class TestReporteAsistenciasExcel:
    @pytest.mark.asyncio
    async def test_admin_downloads_asistencias(self, client, db_session):
        admin = await _create_admin(db_session)
        profesor = await _create_profesor(db_session)
        alumno = await _create_alumno(db_session)
        act = await _create_actividad(db_session)
        turno = await _create_turno(db_session, actividad_id=act.id, profesor_id=profesor.id)
        inscripcion = await _create_inscripcion(db_session, alumno_id=alumno.id, turno_id=turno.id)

        db_session.add(Asistencia(
            inscripcion_id=inscripcion.id, fecha=date(2026, 3, 15), presente=True,
        ))
        db_session.add(Asistencia(
            inscripcion_id=inscripcion.id, fecha=date(2026, 3, 16), presente=False,
        ))
        await db_session.commit()

        r = await client.get(
            "/api/reportes/asistencias/excel",
            params={"fecha_inicio": "2026-03-01", "fecha_fin": "2026-03-31"},
            headers=_auth(admin),
        )
        assert r.status_code == 200
        assert EXCEL_MEDIA in r.headers["content-type"]
        assert "asistencias_" in r.headers["content-disposition"]

        wb = _parse_excel(r.content)
        ws = wb.active
        assert ws.title == "Asistencias"
        assert ws.max_row == 3  # header + 2 records
        assert ws.cell(1, 2).value == "Actividad"

    @pytest.mark.asyncio
    async def test_date_filtering(self, client, db_session):
        admin = await _create_admin(db_session)
        profesor = await _create_profesor(db_session)
        alumno = await _create_alumno(db_session)
        act = await _create_actividad(db_session)
        turno = await _create_turno(db_session, actividad_id=act.id, profesor_id=profesor.id)
        inscripcion = await _create_inscripcion(db_session, alumno_id=alumno.id, turno_id=turno.id)

        db_session.add(Asistencia(
            inscripcion_id=inscripcion.id, fecha=date(2026, 3, 10), presente=True,
        ))
        db_session.add(Asistencia(
            inscripcion_id=inscripcion.id, fecha=date(2026, 3, 20), presente=True,
        ))
        await db_session.commit()

        r = await client.get(
            "/api/reportes/asistencias/excel",
            params={"fecha_inicio": "2026-03-15", "fecha_fin": "2026-03-25"},
            headers=_auth(admin),
        )
        assert r.status_code == 200
        wb = _parse_excel(r.content)
        ws = wb.active
        assert ws.max_row == 2  # header + 1 record (only March 20)

    @pytest.mark.asyncio
    async def test_requires_date_params(self, client, db_session):
        admin = await _create_admin(db_session)
        r = await client.get("/api/reportes/asistencias/excel", headers=_auth(admin))
        assert r.status_code == 422

    @pytest.mark.asyncio
    async def test_non_admin_forbidden(self, client, db_session):
        profesor = await _create_profesor(db_session)
        r = await client.get(
            "/api/reportes/asistencias/excel",
            params={"fecha_inicio": "2026-03-01", "fecha_fin": "2026-03-31"},
            headers=_auth(profesor),
        )
        assert r.status_code == 403


# ── Test: GET /api/reportes/alumno/{id}/pdf ──────────────────────


class TestReporteAlumnoPdf:
    @pytest.mark.asyncio
    async def test_admin_downloads_alumno_pdf(self, client, db_session):
        admin = await _create_admin(db_session)
        alumno = await _create_alumno(db_session)

        r = await client.get(
            f"/api/reportes/alumno/{alumno.id}/pdf", headers=_auth(admin)
        )
        assert r.status_code == 200
        assert PDF_MEDIA in r.headers["content-type"]
        assert "attachment" in r.headers["content-disposition"]
        assert f"alumno_{alumno.id}" in r.headers["content-disposition"]
        # Verify it's a valid PDF
        assert r.content[:4] == b"%PDF"

    @pytest.mark.asyncio
    async def test_pdf_with_full_data(self, client, db_session):
        """PDF with attendance and evaluations."""
        admin = await _create_admin(db_session)
        profesor = await _create_profesor(db_session)
        alumno = await _create_alumno(db_session)
        act = await _create_actividad(db_session)
        turno = await _create_turno(db_session, actividad_id=act.id, profesor_id=profesor.id)
        inscripcion = await _create_inscripcion(db_session, alumno_id=alumno.id, turno_id=turno.id)

        # Add attendance
        db_session.add(Asistencia(
            inscripcion_id=inscripcion.id, fecha=date(2026, 3, 15), presente=True,
        ))
        await db_session.commit()

        # Add evaluations
        await _create_evaluacion(
            db_session, alumno_id=alumno.id, profesional_id=profesor.id,
            peso=80.0, imc=26.1, grasa=22.0,
        )
        await _create_evaluacion(
            db_session, alumno_id=alumno.id, profesional_id=profesor.id,
            peso=77.0, imc=25.1, grasa=20.0,
        )

        r = await client.get(
            f"/api/reportes/alumno/{alumno.id}/pdf", headers=_auth(admin)
        )
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"
        assert len(r.content) > 500  # Should have substantial content

    @pytest.mark.asyncio
    async def test_pdf_alumno_without_data(self, client, db_session):
        """PDF for student with no payments, attendance, or evaluations."""
        admin = await _create_admin(db_session)
        alumno = await _create_alumno(db_session)

        r = await client.get(
            f"/api/reportes/alumno/{alumno.id}/pdf", headers=_auth(admin)
        )
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"

    @pytest.mark.asyncio
    async def test_alumno_not_found(self, client, db_session):
        admin = await _create_admin(db_session)
        r = await client.get("/api/reportes/alumno/9999/pdf", headers=_auth(admin))
        assert r.status_code == 404

    @pytest.mark.asyncio
    async def test_non_alumno_user_returns_404(self, client, db_session):
        """Requesting PDF for a non-student user returns 404."""
        admin = await _create_admin(db_session)
        profesor = await _create_profesor(db_session)

        r = await client.get(
            f"/api/reportes/alumno/{profesor.id}/pdf", headers=_auth(admin)
        )
        assert r.status_code == 404

    @pytest.mark.asyncio
    async def test_non_admin_forbidden(self, client, db_session):
        alumno = await _create_alumno(db_session)
        r = await client.get(
            f"/api/reportes/alumno/{alumno.id}/pdf", headers=_auth(alumno)
        )
        assert r.status_code == 403

    @pytest.mark.asyncio
    async def test_requires_auth(self, client):
        r = await client.get("/api/reportes/alumno/1/pdf")
        assert r.status_code in (401, 403)
