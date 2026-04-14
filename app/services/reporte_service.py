"""Report generation service for Excel and PDF exports."""

import io
from datetime import date, datetime

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.asistencia import Asistencia
from app.models.evaluacion_salud import EvaluacionSalud
from app.models.inscripcion import Inscripcion
from app.models.enums import RolUsuario
from app.models.turno import Turno
from app.models.actividad import Actividad
from app.models.usuario import Usuario


# ── Excel styling constants ─────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HEADER_BORDER = Border(
    bottom=Side(style="thin", color="1A5276"),
    top=Side(style="thin", color="1A5276"),
    left=Side(style="thin", color="1A5276"),
    right=Side(style="thin", color="1A5276"),
)

_CELL_FONT = Font(name="Calibri", size=10)
_CELL_ALIGNMENT = Alignment(vertical="center")
_CELL_BORDER = Border(
    bottom=Side(style="thin", color="D5D8DC"),
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
)
_ALT_ROW_FILL = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")

# PDF color palette
_PDF_PRIMARY = colors.HexColor("#2E86AB")
_PDF_DARK = colors.HexColor("#1A5276")
_PDF_LIGHT_BG = colors.HexColor("#F2F4F4")
_PDF_WHITE = colors.white


def _apply_header_style(ws, row: int, col_count: int) -> None:
    """Apply styled formatting to header row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _HEADER_BORDER


def _apply_data_style(ws, start_row: int, end_row: int, col_count: int) -> None:
    """Apply alternating row styling to data rows."""
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = _CELL_FONT
            cell.alignment = _CELL_ALIGNMENT
            cell.border = _CELL_BORDER
            if (row - start_row) % 2 == 1:
                cell.fill = _ALT_ROW_FILL


def _auto_column_width(ws, col_count: int) -> None:
    """Auto-adjust column widths based on content."""
    for col_idx in range(1, col_count + 1):
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max(max_length + 3, 12), 40
        )


def _workbook_to_bytes(wb: Workbook) -> bytes:
    """Save workbook to bytes buffer."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Excel report generators ─────────────────────────────────────


async def generar_excel_alumnos(db: AsyncSession) -> bytes:
    """Generate Excel with all students and their data."""
    query = (
        select(Usuario)
        .where(Usuario.rol == RolUsuario.ALUMNO)
        .order_by(Usuario.apellido, Usuario.nombre)
    )
    result = await db.execute(query)
    alumnos = list(result.scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Alumnos"

    headers = [
        "ID", "Apellido", "Nombre", "Email", "Telefono",
        "DNI", "Fecha Nacimiento", "Activo", "Fecha Alta",
    ]
    ws.append(headers)
    _apply_header_style(ws, 1, len(headers))

    for alumno in alumnos:
        ws.append([
            alumno.id,
            alumno.apellido,
            alumno.nombre,
            alumno.email,
            alumno.telefono or "",
            alumno.dni or "",
            alumno.fecha_nacimiento.isoformat() if alumno.fecha_nacimiento else "",
            "Si" if alumno.activo else "No",
            alumno.created_at.strftime("%Y-%m-%d") if alumno.created_at else "",
        ])

    if alumnos:
        _apply_data_style(ws, 2, len(alumnos) + 1, len(headers))
    _auto_column_width(ws, len(headers))

    return _workbook_to_bytes(wb)


async def generar_excel_asistencias(
    db: AsyncSession, *, fecha_inicio: date, fecha_fin: date
) -> bytes:
    """Generate Excel with attendance grouped by activity/shift."""
    query = (
        select(Asistencia)
        .join(Inscripcion, Asistencia.inscripcion_id == Inscripcion.id)
        .join(Turno, Inscripcion.turno_id == Turno.id)
        .join(Actividad, Turno.actividad_id == Actividad.id)
        .where(
            Asistencia.fecha >= fecha_inicio,
            Asistencia.fecha <= fecha_fin,
        )
        .options(
            selectinload(Asistencia.inscripcion)
            .selectinload(Inscripcion.turno)
            .selectinload(Turno.actividad),
            selectinload(Asistencia.inscripcion)
            .selectinload(Inscripcion.alumno),
        )
        .order_by(Asistencia.fecha)
    )
    result = await db.execute(query)
    asistencias = list(result.scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    headers = [
        "Fecha", "Actividad", "Turno (Dia/Hora)", "Alumno",
        "DNI", "Presente", "Observacion",
    ]
    ws.append(headers)
    _apply_header_style(ws, 1, len(headers))

    for a in asistencias:
        insc = a.inscripcion
        turno = insc.turno
        actividad = turno.actividad
        alumno = insc.alumno
        ws.append([
            a.fecha.isoformat(),
            actividad.nombre,
            f"{turno.dia_semana.value} {turno.hora_inicio.strftime('%H:%M')}-{turno.hora_fin.strftime('%H:%M')}",
            f"{alumno.apellido}, {alumno.nombre}",
            alumno.dni or "",
            "Si" if a.presente else "No",
            a.observacion or "",
        ])

    if asistencias:
        _apply_data_style(ws, 2, len(asistencias) + 1, len(headers))
    _auto_column_width(ws, len(headers))

    return _workbook_to_bytes(wb)


# ── PDF report generator ────────────────────────────────────────


async def _get_alumno_or_404(db: AsyncSession, alumno_id: int) -> Usuario:
    """Fetch student or raise 404."""
    result = await db.execute(
        select(Usuario).where(Usuario.id == alumno_id)
    )
    alumno = result.scalar_one_or_none()
    if alumno is None or alumno.rol != RolUsuario.ALUMNO:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student not found",
        )
    return alumno


async def _get_asistencias_alumno(db: AsyncSession, alumno_id: int) -> list[Asistencia]:
    """Get attendance records for a student."""
    result = await db.execute(
        select(Asistencia)
        .join(Inscripcion, Asistencia.inscripcion_id == Inscripcion.id)
        .where(Inscripcion.alumno_id == alumno_id)
        .options(
            selectinload(Asistencia.inscripcion)
            .selectinload(Inscripcion.turno)
            .selectinload(Turno.actividad),
        )
        .order_by(Asistencia.fecha.desc())
        .limit(50)
    )
    return list(result.scalars().all())


async def _get_evaluaciones_alumno(
    db: AsyncSession, alumno_id: int
) -> list[EvaluacionSalud]:
    """Get health evaluations for a student, ordered by date."""
    result = await db.execute(
        select(EvaluacionSalud)
        .where(EvaluacionSalud.alumno_id == alumno_id)
        .options(selectinload(EvaluacionSalud.profesional))
        .order_by(EvaluacionSalud.fecha)
    )
    return list(result.scalars().all())


def _build_pdf_header_style() -> ParagraphStyle:
    """Create PDF header paragraph style."""
    return ParagraphStyle(
        "CustomTitle",
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=_PDF_DARK,
        spaceAfter=6 * mm,
    )


def _build_section_style() -> ParagraphStyle:
    """Create PDF section title style."""
    return ParagraphStyle(
        "SectionTitle",
        fontName="Helvetica-Bold",
        fontSize=13,
        textColor=_PDF_PRIMARY,
        spaceBefore=8 * mm,
        spaceAfter=4 * mm,
    )


def _build_body_style() -> ParagraphStyle:
    """Create PDF body text style."""
    styles = getSampleStyleSheet()
    return ParagraphStyle(
        "CustomBody",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        spaceAfter=2 * mm,
    )


def _make_table(data: list[list], col_widths: list[float] | None = None) -> Table:
    """Create a styled PDF table."""
    table = Table(data, colWidths=col_widths, repeatRows=1)
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), _PDF_PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), _PDF_WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D5D8DC")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    # Alternating row colors
    for i in range(1, len(data)):
        if i % 2 == 0:
            style_commands.append(
                ("BACKGROUND", (0, i), (-1, i), _PDF_LIGHT_BG)
            )
    table.setStyle(TableStyle(style_commands))
    return table


def _build_evolution_chart(evaluaciones: list[EvaluacionSalud], page_width: float) -> Table:
    """Build a simple text-based evolution table for health metrics."""
    if not evaluaciones:
        return None

    headers = ["Fecha", "Peso (kg)", "Altura (cm)", "IMC", "Grasa Corp. (%)"]
    data = [headers]
    for ev in evaluaciones:
        fecha_str = ev.fecha.strftime("%Y-%m-%d") if ev.fecha else ""
        data.append([
            fecha_str,
            f"{float(ev.peso_kg):.1f}" if ev.peso_kg else "-",
            f"{float(ev.altura_cm):.1f}" if ev.altura_cm else "-",
            f"{float(ev.imc):.1f}" if ev.imc else "-",
            f"{float(ev.grasa_corporal):.1f}" if ev.grasa_corporal else "-",
        ])

    col_w = page_width / 5
    return _make_table(data, [col_w] * 5)


def _build_evolution_graph_drawing(evaluaciones: list[EvaluacionSalud], width: float) -> Table | None:
    """Build a simple ASCII-style evolution indicator using arrows."""
    if len(evaluaciones) < 2:
        return None

    metrics = []
    last = evaluaciones[-1]
    prev = evaluaciones[-2]

    def _trend(current, previous, label, unit, lower_better=False):
        if current is None or previous is None:
            return None
        curr_f = float(current)
        prev_f = float(previous)
        diff = curr_f - prev_f
        if abs(diff) < 0.01:
            arrow = "="
        elif (diff < 0 and lower_better) or (diff > 0 and not lower_better):
            arrow = "+"
        else:
            arrow = "-"
        return [label, f"{prev_f:.1f}", f"{curr_f:.1f}", f"{diff:+.1f} {unit}", arrow]

    peso_row = _trend(last.peso_kg, prev.peso_kg, "Peso", "kg", lower_better=True)
    imc_row = _trend(last.imc, prev.imc, "IMC", "", lower_better=True)
    grasa_row = _trend(last.grasa_corporal, prev.grasa_corporal, "Grasa Corp.", "%", lower_better=True)

    for row in [peso_row, imc_row, grasa_row]:
        if row:
            metrics.append(row)

    if not metrics:
        return None

    data = [["Metrica", "Anterior", "Actual", "Diferencia", "Tendencia"]] + metrics
    col_w = width / 5
    return _make_table(data, [col_w] * 5)


async def generar_pdf_alumno(db: AsyncSession, *, alumno_id: int) -> bytes:
    """Generate a professional PDF profile card for a student."""
    alumno = await _get_alumno_or_404(db, alumno_id)
    asistencias = await _get_asistencias_alumno(db, alumno_id)
    evaluaciones = await _get_evaluaciones_alumno(db, alumno_id)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
    )
    page_width = A4[0] - 4 * cm

    title_style = _build_pdf_header_style()
    section_style = _build_section_style()
    body_style = _build_body_style()

    elements = []

    # Title
    elements.append(Paragraph("HSP-70 GESTION", title_style))
    elements.append(Paragraph("Ficha del Alumno", section_style))
    elements.append(Spacer(1, 4 * mm))

    # Personal data
    elements.append(Paragraph("Datos Personales", section_style))
    personal_data = [
        ["Campo", "Valor"],
        ["Nombre", f"{alumno.nombre} {alumno.apellido}"],
        ["Email", alumno.email],
        ["Telefono", alumno.telefono or "-"],
        ["DNI", alumno.dni or "-"],
        ["Fecha Nacimiento", alumno.fecha_nacimiento.isoformat() if alumno.fecha_nacimiento else "-"],
        ["Estado", "Activo" if alumno.activo else "Inactivo"],
        ["Fecha Alta", alumno.created_at.strftime("%Y-%m-%d") if alumno.created_at else "-"],
    ]
    elements.append(_make_table(personal_data, [page_width * 0.35, page_width * 0.65]))

    # Attendance history
    elements.append(Paragraph("Historial de Asistencias", section_style))
    if asistencias:
        asist_headers = ["Fecha", "Actividad", "Turno", "Presente"]
        asist_data = [asist_headers]
        for a in asistencias[:30]:
            turno = a.inscripcion.turno
            actividad = turno.actividad
            asist_data.append([
                a.fecha.isoformat(),
                actividad.nombre,
                f"{turno.dia_semana.value} {turno.hora_inicio.strftime('%H:%M')}",
                "Si" if a.presente else "No",
            ])
        col_w = page_width / 4
        elements.append(_make_table(asist_data, [col_w] * 4))

        total = len(asistencias)
        presentes = sum(1 for a in asistencias if a.presente)
        pct = (presentes / total * 100) if total > 0 else 0
        elements.append(Spacer(1, 2 * mm))
        elements.append(
            Paragraph(
                f"Asistencia: {presentes}/{total} ({pct:.0f}%)",
                body_style,
            )
        )
    else:
        elements.append(Paragraph("Sin asistencias registradas.", body_style))

    # Health evaluations
    elements.append(Paragraph("Evaluaciones de Salud", section_style))
    if evaluaciones:
        ev_table = _build_evolution_chart(evaluaciones, page_width)
        if ev_table:
            elements.append(ev_table)

        # Evolution trends
        trend_table = _build_evolution_graph_drawing(evaluaciones, page_width)
        if trend_table:
            elements.append(Spacer(1, 4 * mm))
            elements.append(Paragraph("Evolucion", section_style))
            elements.append(trend_table)

        # Notes from latest evaluation
        latest = evaluaciones[-1]
        if latest.objetivo or latest.notas:
            elements.append(Spacer(1, 3 * mm))
            if latest.objetivo:
                elements.append(
                    Paragraph(f"<b>Objetivo:</b> {latest.objetivo}", body_style)
                )
            if latest.notas:
                elements.append(
                    Paragraph(f"<b>Notas:</b> {latest.notas}", body_style)
                )
    else:
        elements.append(Paragraph("Sin evaluaciones de salud.", body_style))

    # Footer with generation date
    elements.append(Spacer(1, 10 * mm))
    footer_style = ParagraphStyle(
        "Footer",
        fontName="Helvetica",
        fontSize=8,
        textColor=colors.HexColor("#999999"),
    )
    elements.append(
        Paragraph(
            f"Generado el {datetime.now().strftime('%Y-%m-%d %H:%M')} - HSP-70 Gestion",
            footer_style,
        )
    )

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()
